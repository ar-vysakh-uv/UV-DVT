#Author - Dhananjay

import can
import random
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from scram_implementation import scram_decode 


# === CONFIGURATION ===
CHANNEL = 'PCAN_USBBUS1'
BITRATE = 500000
START = 0xE157863
START_DATA = [0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00, 0x00]
WAIT_FOR_ID = 0x176
CHALLENGE_ID = 0xAC11111
RESPONSE_ID = 0xAC11117  # Extended ID
RESULT_ID = 0xAC1111F    # Extended ID
EXCEL_FILE = 'scram_log.xlsx'
DATA = [0x3A, 0x32, 0x4C, 0xEB, 0x1E, 0x52, 0xFC, 0x4B]
KEY = [0x3A, 0x7F, 0xC4, 0xE9, 0x5B, 0xA2, 0x18, 0xD7]


def send_and_wait(bus):
    try:
        
            response = bus.recv(timeout=5)  # timeout in seconds

        #  Step 1: Wait for the specific response from DCP
            print(f"[INFO] Waiting for DCP wake up message from ID: 0x{WAIT_FOR_ID:X}...")
            while True:
                response = bus.recv(timeout=1)
                if response is None:
                    continue  # Nothing received, keep waiting
                if response.arbitration_id == WAIT_FOR_ID:
                    print(f"[SUCCESS] Received expected message: {' '.join(f'{b:02X}' for b in response.data)}")
                    break  # Exit loop and continue processing
                else:
                    print(f"[INFO] Ignored ID: 0x{response.arbitration_id:X}")

            # Step 2 : Send your charger's first message as all ZEROs
            msg = can.Message(arbitration_id=START, data= START_DATA, is_extended_id=True)
            bus.send(msg)
            print(f"[INFO] Sent: {' '.join(f'{b:02X}' for b in START_DATA)}")

            # Step 3: Wait for CHALLENGE_ID response
            print(f"[INFO] Waiting for challenge from ID: 0x{CHALLENGE_ID:X}...")
            while True:
                challenge_msg = bus.recv(timeout=3)
                if challenge_msg is None:
                    print("[WARN] Timeout waiting for challenge response.")
                    break
                if challenge_msg.arbitration_id == CHALLENGE_ID:
                    print(f"[SUCCESS] Received Challenge: {' '.join(f'{b:02X}' for b in challenge_msg.data)}")
                    break
                else:
                    print(f"[INFO] Ignored ID: {' '.join(f'{b:02X}' for b in challenge_msg.data)}")
            
            challenge = list(challenge_msg.data)
            
            print("Challenge (raw):", ' '.join(f'{b:02X}' for b in challenge))
            return challenge

    except can.CanError as e:
            print(f"[ERROR] CAN error: {e}")



def send_response_and_wait(bus,response):
    try:
        
            msg = can.Message(arbitration_id=RESPONSE_ID, data= response, is_extended_id=True)
            bus.send(msg)
            print(f"[INFO] Sent: {' '.join(f'{b:02X}' for b in response)}")
        
            print(f"[INFO] Waiting for response from AC1111F... ")

            ack = bus.recv(timeout=5)  # timeout in seconds

            if ack:
                if ack.arbitration_id == RESULT_ID:
                    print(f"[INFO] Received from {hex(ack.arbitration_id)}: {' '.join(f'{b:02X}' for b in ack.data)}")
                else:
                    print(f"[WARNING] Got data from unexpected ID: {hex(ack.arbitration_id)}")

    except can.CanError as e:
        print(f"[ERROR] CAN error: {e}")


def log_to_excel(file_name, challenge, response, ack_data):
    if os.path.exists(file_name):
        wb = load_workbook(file_name)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Timestamp", "Challenge", "Response", "ACK Message"])

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Format data
    challenge_str = ' '.join(f'{b:02X}' for b in challenge)
    response_str = ' '.join(f'{b:02X}' for b in response)
    ack_str = ' '.join(f'{b:02X}' for b in ack_data) if ack_data else "Timeout or Invalid"

    ws.append([timestamp, challenge_str, response_str, ack_str])
    wb.save(file_name)

if __name__ == "__main__":
# while(1):

    with can.interface.Bus(channel=CHANNEL, interface='pcan', bitrate=BITRATE) as bus:
        challenge = send_and_wait(bus)
        if challenge is not None:
            response = scram_decode(challenge)
            if response:
                try:
                    msg = can.Message(arbitration_id=RESPONSE_ID, data=response, is_extended_id=True)
                    bus.send(msg)
                    print(f"[INFO] Sent: {' '.join(f'{b:02X}' for b in response)}")

                    print(f"[INFO] Waiting for response from AC1111F...")
                    ack = bus.recv(timeout=5)

                    if ack and ack.arbitration_id == RESULT_ID:
                        print(f"[INFO] Received from {hex(ack.arbitration_id)}: {' '.join(f'{b:02X}' for b in ack.data)}")
                        ack_data = ack.data
                    else:
                        print("[WARNING] No ACK or unexpected ID.")
                        ack_data = None

                    # Log the data
                    log_to_excel(EXCEL_FILE, challenge, response, ack_data)

                except can.CanError as e:
                    print(f"[ERROR] CAN error: {e}")
        else:
            print("[ERROR] Did not receive a valid challenge.")